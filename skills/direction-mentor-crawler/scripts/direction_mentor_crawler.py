#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urldefrag, urljoin, urlparse

import requests
from bs4 import BeautifulSoup


EMAIL_RE = re.compile(r"[A-Za-z0-9_.+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
EMAIL_AT_RE = re.compile(r"([A-Za-z0-9_.+-]+)\s*(?:\[at\]|\(at\)| at )\s*([A-Za-z0-9.-]+\.[A-Za-z]{2,})", re.I)
PHONE_RE = re.compile(r"(?:电话|办公电话|Tel|Phone|联系方式)[:：\s]*([0-9+\-()（）\s]{7,})", re.I)
TITLE_RE = re.compile(r"(院士|教授|副教授|助理教授|研究员|副研究员|高级工程师|讲师|博导|硕导|博士生导师|硕士生导师)")
TITLE_PREFIXES = ("长聘", "预聘", "正高级", "副高级", "助理", "特聘")
CHINESE_NAME_RE = re.compile(r"^[\u4e00-\u9fff·]{2,5}$")
JS_REDIRECT_RE = re.compile(r"window\.location\.href\s*=\s*['\"]([^'\"]+)['\"]", re.I)

NOISE = {
    "首页",
    "学院概况",
    "师资队伍",
    "专职教师",
    "院士风采",
    "客座教授",
    "博士后",
    "光荣退休",
    "科学研究",
    "人才培养",
    "平台基地",
    "学生工作",
    "招贤纳士",
    "校友中心",
    "联系我们",
    "English",
    "北大主页",
}


@dataclass
class TeacherLink:
    name_hint: str
    title_hint: str
    url: str


@dataclass
class TeacherTableRow:
    name: str
    title: str
    research: str
    homepage: str
    source_url: str


@dataclass
class Page:
    url: str
    html: str
    soup: BeautifulSoup
    text: str
    title: str


def fetch(session: requests.Session, url: str) -> Page:
    resp = session.get(url, timeout=15)
    resp.raise_for_status()
    resp.encoding = resp.apparent_encoding or resp.encoding
    soup = BeautifulSoup(resp.text, "html.parser")
    title = soup.title.get_text(" ", strip=True) if soup.title else ""
    clean = BeautifulSoup(resp.text, "html.parser")
    for tag in clean(["script", "style", "noscript"]):
        tag.decompose()
    text = re.sub(r"\s+", " ", clean.get_text(" ", strip=True))
    return Page(url=url, html=resp.text, soup=soup, text=text, title=title)


def normalize(url: str) -> str:
    return urldefrag(url.strip())[0]


def same_netloc(url: str, netloc: str) -> bool:
    parsed = urlparse(url)
    return parsed.scheme in {"http", "https"} and parsed.netloc == netloc


def split_name_title(text: str) -> tuple[str, str]:
    text = re.sub(r"\s+", " ", text).strip()
    if not text or text in NOISE:
        return "", ""
    compact = text.replace(" ", "")
    title_match = TITLE_RE.search(compact)
    if title_match:
        start = title_match.start()
        for prefix in TITLE_PREFIXES:
            prefix_start = start - len(prefix)
            if prefix_start >= 2 and compact[prefix_start:start] == prefix:
                start = prefix_start
                break
        name = compact[:start]
        title = compact[start:].strip(" ，,;；")
    else:
        name = compact
        title = ""
    if not CHINESE_NAME_RE.match(name) or name in NOISE or "教师" in name:
        return "", ""
    return name, title


def is_detail_href(href: str, list_url: str, name: str) -> bool:
    parsed = urlparse(href)
    list_path = urlparse(list_url).path
    base_dir = list_path.rsplit("/", 1)[0] + "/"
    if parsed.path.startswith(base_dir) and parsed.path != list_path and parsed.path.endswith("/index.htm"):
        return True
    if name and re.search(r"/info/\d+/\d+\.htm$", parsed.path):
        return True
    return False


def extract_teacher_links(list_page: Page) -> list[TeacherLink]:
    netloc = urlparse(list_page.url).netloc
    links: dict[str, TeacherLink] = {}

    for a in list_page.soup.find_all("a", href=True):
        url = normalize(urljoin(list_page.url, a["href"]))
        raw_text = a.get_text(" ", strip=True)
        name, title = split_name_title(raw_text)
        if not name:
            continue
        if not same_netloc(url, netloc) or not is_detail_href(url, list_page.url, name):
            continue
        links[url] = TeacherLink(name, title, url)

    return list(links.values())


def header_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    for idx, header in enumerate(headers):
        normalized = re.sub(r"\s+", "", header)
        if any(candidate in normalized for candidate in candidates):
            return idx
    return None


def extract_teacher_table_rows(list_page: Page) -> list[TeacherTableRow]:
    """Extract mentor rows from official table-style pages.

    Some authoritative sources, especially graduate admission catalogs, list
    mentors directly in tables without linking to individual detail pages. In
    that case the table itself is the coverage unit.
    """
    rows: list[TeacherTableRow] = []
    for table in list_page.soup.find_all("table"):
        trs = table.find_all("tr")
        if len(trs) < 2:
            continue
        headers = [cell.get_text(" ", strip=True) for cell in trs[0].find_all(["th", "td"])]
        if not headers:
            continue
        name_idx = header_index(headers, ("姓名", "教师", "导师", "名称"))
        title_idx = header_index(headers, ("职称", "身份", "职务"))
        research_idx = header_index(headers, ("研究方向", "方向", "招生方向", "备注"))
        link_idx = header_index(headers, ("链接", "主页", "介绍"))
        if name_idx is None:
            continue
        for tr in trs[1:]:
            cells = tr.find_all(["td", "th"])
            if len(cells) <= name_idx:
                continue
            values = [cell.get_text(" ", strip=True) for cell in cells]
            name = re.sub(r"\s+", " ", values[name_idx]).strip()
            if not name or name in NOISE:
                continue
            title = values[title_idx] if title_idx is not None and title_idx < len(values) else ""
            research = values[research_idx] if research_idx is not None and research_idx < len(values) else ""
            homepage = values[link_idx] if link_idx is not None and link_idx < len(values) else ""
            if link_idx is not None and link_idx < len(cells):
                a = cells[link_idx].find("a", href=True)
                if a:
                    homepage = normalize(urljoin(list_page.url, a["href"]))
            rows.append(TeacherTableRow(name=name, title=title, research=research, homepage=homepage, source_url=list_page.url))
    return rows


def extract_catalog_section_rows(list_page: Page, section_keyword: str) -> list[TeacherTableRow]:
    """Extract rows from large admission catalogs by program/major heading.

    Tsinghua-style catalogs often contain one huge table where a major heading
    row is followed by direction rows. This parser starts at the row containing
    `section_keyword` and stops at the next major-code row.
    """
    rows: list[TeacherTableRow] = []
    if not section_keyword:
        return rows
    major_code_re = re.compile(r"^\s*\d{4}[A-Z]?\d*")
    direction_re = re.compile(r"^\s*\d{2}[（(]")
    for table in list_page.soup.find_all("table"):
        trs = table.find_all("tr")
        in_section = False
        current_direction = ""
        for tr in trs:
            values = [re.sub(r"\s+", " ", cell.get_text(" ", strip=True)).strip() for cell in tr.find_all(["td", "th"])]
            values = [v for v in values if v]
            if not values:
                continue
            joined = " ".join(values)
            if not in_section and section_keyword in joined:
                in_section = True
                continue
            if not in_section:
                continue
            if section_keyword not in joined and major_code_re.match(values[0]) and not direction_re.match(values[0]):
                break
            if direction_re.match(values[0]):
                current_direction = values[0]
                name_values = values[1:]
            else:
                name_values = values
            for value in name_values:
                if not value or value in {"不可推免", "仅推免"}:
                    continue
                if direction_re.match(value) or major_code_re.match(value):
                    continue
                if len(value) > 40:
                    continue
                rows.append(
                    TeacherTableRow(
                        name=value,
                        title="",
                        research=current_direction,
                        homepage="",
                        source_url=list_page.url,
                    )
                )
    unique: dict[tuple[str, str], TeacherTableRow] = {}
    for row in rows:
        unique[(row.name, row.research)] = row
    return list(unique.values())


def table_rows_to_records(
    rows: list[TeacherTableRow],
    school: str,
    college: str,
    direction: str,
) -> list[dict[str, str]]:
    return [
        {
            "学校": school,
            "学院": college,
            "方向/系": direction,
            "姓名": row.name,
            "职称/身份": row.title,
            "研究方向": row.research,
            "邮箱": "",
            "电话": "",
            "主页/实验室链接": row.homepage,
            "详情页": row.source_url,
        }
        for row in rows
    ]


def label_map(page: Page) -> dict[str, str]:
    result: dict[str, str] = {}
    for node in page.soup.select(".name"):
        label = node.get_text(" ", strip=True).strip(" ：:")
        sibling = node.find_next_sibling(class_="text")
        if sibling:
            result[label] = sibling.get_text(" ", strip=True)
    return result


def likely_name(value: str) -> str:
    value = re.sub(r"\s+", "", value)
    value = re.sub(r"(教授|老师|导师|博士|主页|简介|个人信息)$", "", value)
    if CHINESE_NAME_RE.match(value) and value not in NOISE and "教师" not in value:
        return value
    return ""


def extract_name(page: Page, hint: str) -> str:
    labels = label_map(page)
    for candidate in (
        labels.get("姓名", ""),
        hint,
        page.title.split("-")[0],
        page.title.split("_")[0],
    ):
        name = likely_name(candidate)
        if name:
            return name
    for selector in ("h1", "h2", ".title", ".tit", ".teacher-name"):
        for node in page.soup.select(selector):
            name = likely_name(node.get_text(" ", strip=True))
            if name:
                return name
    return ""


def extract_links(page: Page, include_external: bool) -> str:
    netloc = urlparse(page.url).netloc
    hits: list[str] = []
    for a in page.soup.find_all("a", href=True):
        text = a.get_text(" ", strip=True).lower()
        href = normalize(urljoin(page.url, a["href"]))
        haystack = f"{text} {href}".lower()
        if not any(k in haystack for k in ("课题组", "实验室", "个人主页", "主页", "lab", "group", "homepage", "scholar")):
            continue
        if href.startswith("mailto:"):
            continue
        if not include_external and urlparse(href).netloc not in {"", netloc}:
            continue
        if href not in hits:
            hits.append(href)
    return "；".join(hits[:8])


def first_label(labels: dict[str, str], names: tuple[str, ...]) -> str:
    for name in names:
        if labels.get(name):
            return labels[name]
    return ""


def extract_record(
    page: Page,
    teacher: TeacherLink,
    school: str,
    college: str,
    direction: str,
    include_external_homepages: bool,
) -> dict[str, str]:
    labels = label_map(page)
    emails = set(EMAIL_RE.findall(page.text))
    emails.update(f"{left}@{right}" for left, right in EMAIL_AT_RE.findall(page.text))
    phone = PHONE_RE.search(page.text)
    name = extract_name(page, teacher.name_hint)
    title = first_label(labels, ("职称", "职务")) or teacher.title_hint
    if not title:
        found_titles = []
        for item in TITLE_RE.findall(page.text[:1200]):
            if item not in found_titles:
                found_titles.append(item)
        title = "；".join(found_titles[:5])
    research = first_label(labels, ("研究领域", "研究方向", "主要研究方向", "招生方向"))
    homepage = first_label(labels, ("课题组网站", "个人主页", "主页")) or extract_links(page, include_external_homepages)
    return {
        "学校": school,
        "学院": college,
        "方向/系": direction,
        "姓名": name,
        "职称/身份": title,
        "研究方向": research,
        "邮箱": "；".join(sorted(emails)),
        "电话": phone.group(1).strip() if phone else "",
        "主页/实验室链接": homepage,
        "详情页": page.url,
    }


def write_csv(path: Path, rows: list[dict[str, str]], headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx_if_available(path: Path, rows: list[dict[str, str]], headers: list[str]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "mentors"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")
    sheet.freeze_panes = "A2"
    for idx, header in enumerate(headers, start=1):
        width = max(len(str(header)) + 2, 14)
        for row in rows[:200]:
            width = max(width, min(len(str(row.get(header, ""))) + 2, 45))
        sheet.column_dimensions[get_column_letter(idx)].width = width
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="方向级导师信息抓取：从一个已确认的师资列表页抓详情。")
    parser.add_argument("--list-url", required=True)
    parser.add_argument("--school", default="")
    parser.add_argument("--college", default="")
    parser.add_argument("--direction", default="")
    parser.add_argument("--out-dir", default="skill-outputs/01-direction-mentor-crawler/YYYY-MM-DD-school-college-direction-mentors")
    parser.add_argument("--delay", type=float, default=0.2)
    parser.add_argument("--include-external-homepages", action="store_true")
    parser.add_argument("--debug-html", action="store_true")
    parser.add_argument(
        "--source-type",
        choices=("auto", "detail-links", "table"),
        default="auto",
        help="auto: prefer detail links and fall back to official tables; detail-links: require teacher detail pages; table: parse official table rows.",
    )
    parser.add_argument(
        "--section-keyword",
        default="",
        help="For large admission catalogs, parse only the table section after this program/major keyword, stopping at the next major heading.",
    )
    args = parser.parse_args()

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 direction-mentor-crawler/1.0"})
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    list_page = fetch(session, args.list_url)
    redirects = list(JS_REDIRECT_RE.findall(list_page.html))
    if redirects:
        redirected_url = normalize(urljoin(list_page.url, redirects[0]))
        if same_netloc(redirected_url, urlparse(list_page.url).netloc):
            list_page = fetch(session, redirected_url)
    teacher_links = [] if args.source_type == "table" else extract_teacher_links(list_page)
    if not teacher_links and args.source_type in {"auto", "table"}:
        table_rows = extract_catalog_section_rows(list_page, args.section_keyword) if args.section_keyword else extract_teacher_table_rows(list_page)
        if table_rows:
            records = table_rows_to_records(table_rows, args.school, args.college, args.direction)
            checks = [
                {
                    "列表页姓名": row.name,
                    "列表页职称": row.title,
                    "详情页": row.source_url,
                    "抓取状态": "ok",
                    "错误": "official_table_source",
                }
                for row in table_rows
            ]
            headers = ["学校", "学院", "方向/系", "姓名", "职称/身份", "研究方向", "邮箱", "电话", "主页/实验室链接", "详情页"]
            write_csv(out_dir / "mentors.csv", records, headers)
            write_xlsx_if_available(out_dir / "mentors.xlsx", records, headers)
            write_csv(out_dir / "coverage_check.csv", checks, ["列表页姓名", "列表页职称", "详情页", "抓取状态", "错误"])
            with (out_dir / "mentors.jsonl").open("w", encoding="utf-8") as f:
                for record in records:
                    f.write(json.dumps(record, ensure_ascii=False) + "\n")
            (out_dir / "source_notes.md").write_text(
                "本次使用官方表格源抽取导师信息。该页面没有可稳定识别的教师详情链接，覆盖范围以该官方表格为准。\n",
                encoding="utf-8",
            )
            print(f"官方表格导师行：{len(table_rows)}；输出目录：{out_dir.resolve()}")
            return
    if not teacher_links:
        raise SystemExit("未在列表页识别到教师详情链接或导师表格；请确认传入的是方向/系师资页、招生目录页或项目导师表。")

    records: list[dict[str, str]] = []
    checks: list[dict[str, str]] = []
    debug_dir = out_dir / "debug_html"
    if args.debug_html:
        debug_dir.mkdir(exist_ok=True)
        (debug_dir / "list_page.html").write_text(list_page.html, encoding="utf-8")

    for idx, teacher in enumerate(teacher_links, 1):
        status = "ok"
        error = ""
        try:
            page = fetch(session, teacher.url)
            if args.debug_html:
                safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", teacher.name_hint or str(idx))
                (debug_dir / f"{idx:03d}_{safe}.html").write_text(page.html, encoding="utf-8")
            record = extract_record(
                page,
                teacher,
                args.school,
                args.college,
                args.direction,
                args.include_external_homepages,
            )
            if not record["姓名"]:
                record["姓名"] = teacher.name_hint
            if not record["姓名"]:
                status = "missing_name"
            records.append(record)
        except Exception as exc:  # noqa: BLE001 - coverage file should keep failures visible.
            status = "failed"
            error = repr(exc)
        checks.append(
            {
                "列表页姓名": teacher.name_hint,
                "列表页职称": teacher.title_hint,
                "详情页": teacher.url,
                "抓取状态": status,
                "错误": error,
            }
        )
        print(f"{idx}/{len(teacher_links)} {teacher.name_hint or teacher.url} {status}")
        if args.delay:
            time.sleep(args.delay)

    headers = ["学校", "学院", "方向/系", "姓名", "职称/身份", "研究方向", "邮箱", "电话", "主页/实验室链接", "详情页"]
    write_csv(out_dir / "mentors.csv", records, headers)
    write_xlsx_if_available(out_dir / "mentors.xlsx", records, headers)
    write_csv(out_dir / "coverage_check.csv", checks, ["列表页姓名", "列表页职称", "详情页", "抓取状态", "错误"])
    with (out_dir / "mentors.jsonl").open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    ok = sum(1 for row in checks if row["抓取状态"] == "ok")
    print(f"列表页教师链接：{len(teacher_links)}；成功：{ok}；输出目录：{out_dir.resolve()}")


if __name__ == "__main__":
    main()
